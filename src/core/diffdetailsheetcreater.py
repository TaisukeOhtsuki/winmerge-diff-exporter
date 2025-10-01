# -*- coding: UTF-8 -*-
"""
Diff detail sheet creator module
"""
from typing import List, Tuple
from openpyxl.styles import Font, Color, PatternFill
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .config import config
from .utils import ExcelFormatter
from .common import Timer, logger

timer_DDSC = Timer("DiffDetailSheetCreator")

class DiffDetailSheetCreator:
    """Creates detailed diff sheets from Excel workbook"""
    
    def __init__(self, file_path: str, start_index: int = None, context_lines: int = None, sheet_name_to_filename: dict = None):
        self.file_path = file_path
        self.wb = load_workbook(self.file_path)
        self.start_index = start_index or config.diff.sheet_start_index
        self.context_lines = context_lines or config.diff.context_lines
        self.detail_ws = self.wb.create_sheet(index=0, title='compare')
        self.row_cursor = 2
        self.sheet_name_to_filename = sheet_name_to_filename or {}
        
        logger.info(f"DiffDetailSheetCreator initialized for: {file_path}")

    def generate(self) -> None:
        """Generate detailed diff sheet"""
        timer_DDSC.start(memo="generate")
        
        try:
            worksheets_to_process = self.wb.worksheets[self.start_index:]
            logger.info(f"Processing {len(worksheets_to_process)} worksheets for compare sheet")
            
            for ws in worksheets_to_process:
                logger.info(f"[DEBUG] Processing sheet: '{ws.title}'")
                file_name = self._extract_filename(ws.title)
                logger.info(f"[DEBUG] Extracted filename: '{file_name}'")
                self._write_filename_label(file_name)
                self._process_sheet(ws)
            
            ExcelFormatter.set_worksheet_format(self.detail_ws)
            self.wb.save(self.file_path)
            logger.info("Diff detail sheet generation completed")
            
        except Exception as e:
            logger.error(f"Failed to generate diff detail sheet: {e}")
            raise
        finally:
            timer_DDSC.stop()

    def _extract_filename(self, sheet_name: str) -> str:
        """Extract filename from sheet name using mapping or fallback logic"""
        logger.info(f"[DEBUG] _extract_filename input: '{sheet_name}'")
        
        # First, try to use the mapping provided by WinMergeXlsx
        if sheet_name in self.sheet_name_to_filename:
            filename = self.sheet_name_to_filename[sheet_name]
            logger.info(f"[DEBUG] Found in mapping: '{sheet_name}' -> '{filename}'")
            return filename
        
        # Fallback: Extract from sheet name
        # Sheet name format: folder1_folder2_..._filename
        # Convert underscores to backslashes temporarily
        path_like = sheet_name.replace('_', '\\')
        logger.info(f"[DEBUG] After underscore replacement: '{path_like}'")
        
        # Split by backslash and get the last component (filename)
        if '\\' in path_like:
            result = path_like.split('\\')[-1]
            logger.info(f"[DEBUG] Extracted last component: '{result}'")
            return result
        
        # If no backslash, return as-is (it's already just a filename)
        logger.info(f"[DEBUG] No backslash found, returning as-is: '{sheet_name}'")
        return sheet_name

    def _write_filename_label(self, file_name: str) -> None:
        """Write filename label in the detail sheet"""
        cell = self.detail_ws.cell(row=self.row_cursor, column=1)
        cell.value = file_name
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')
        self.row_cursor += 1

    def _process_sheet(self, ws) -> None:
        """Process individual worksheet for diff detection"""
        max_row = self._get_max_colored_row(ws)
        diff_rows = self._detect_diff_rows(ws, max_row)
        blocks = self._merge_diff_blocks(diff_rows)

        logger.info(f"Found {len(diff_rows)} diff rows in {len(blocks)} blocks for sheet: {ws.title}")

        for block_start, block_end in blocks:
            self._copy_block(ws, block_start, block_end)
            self.row_cursor += 2  # Add spacing between blocks

        self.row_cursor += 2  # Add spacing after processing each sheet

    def _get_max_colored_row(self, ws) -> int:
        """Find the maximum row with colored cells"""
        for row in reversed(range(1, ws.max_row + 1)):
            cell = ws.cell(row=row, column=1)
            if cell.fill.start_color.rgb not in ('FFFFFFFF', '00000000'):
                return row
        
        logger.warning(f"No colored rows found in sheet: {ws.title}, using max_row")
        return ws.max_row

    def _detect_diff_rows(self, ws, max_row: int) -> List[int]:
        """Detect rows containing differences"""
        timer_DDSC.start(memo="detect_diff_rows")
        
        diff_rows = set()
        yellow_color = config.diff.yellow_color
        
        # Track unique colors for debugging
        unique_colors = set()
        
        for row in range(config.excel.diff_start_row, max_row + 1):
            for col_config in config.diff_formats['code']:
                col_letter = col_config['col']
                col = column_index_from_string(col_letter)
                cell = ws.cell(row=row, column=col)
                
                # Collect color information for debugging
                cell_color = cell.fill.start_color.rgb
                if cell_color and cell_color not in ('FFFFFFFF', '00000000', None):
                    unique_colors.add(cell_color)
                
                if cell_color == yellow_color:
                    diff_rows.add(row)
                    break
        
        # Log detected colors for debugging
        if unique_colors:
            logger.info(f"[DEBUG] Sheet '{ws.title}' - Unique colors found: {unique_colors}")
            logger.info(f"[DEBUG] Looking for color: '{yellow_color}'")
        
        timer_DDSC.stop()
        return sorted(diff_rows)

    def _merge_diff_blocks(self, diff_rows: List[int]) -> List[Tuple[int, int]]:
        """Merge adjacent diff rows into blocks with context"""
        timer_DDSC.start(memo="merge_diff_blocks")
        
        if not diff_rows:
            return []

        # Create ranges with context
        ranges = []
        for row in diff_rows:
            start = max(row - self.context_lines, config.excel.diff_start_row)
            end = row + self.context_lines
            ranges.append((start, end))

        # Merge overlapping ranges
        merged = []
        current_start, current_end = ranges[0]

        for start, end in ranges[1:]:
            if start <= current_end + 1:
                current_end = max(current_end, end)
            else:
                merged.append((current_start, current_end))
                current_start, current_end = start, end

        merged.append((current_start, current_end))
        
        timer_DDSC.stop()
        return merged

    def _copy_block(self, ws, start_row: int, end_row: int) -> None:
        """Copy a block of rows from source worksheet to detail sheet"""
        for row in range(start_row, end_row + 1):
            # Copy columns 1-4 to positions 1-4 in detail sheet
            for src_col, dst_col in zip([1, 2, 3, 4], range(1, 5)):
                source = ws.cell(row=row, column=src_col)
                target = self.detail_ws.cell(row=self.row_cursor, column=dst_col)
                
                # Copy value
                target.value = source.value

                # Copy font with color
                if source.font and source.font.color:
                    font_color = getattr(source.font.color, 'rgb', None)
                    if font_color and isinstance(font_color, str) and font_color not in ('00000000', '0'):
                        target.font = Font(color=Color(rgb=font_color))
                    else:
                        target.font = Font()
                else:
                    target.font = Font()

                # Copy fill color (only if it's actually colored, not white or black)
                if source.fill and source.fill.start_color:
                    fill_color = getattr(source.fill.start_color, 'rgb', None)
                    # Check if fill color is valid and not default colors
                    if fill_color and isinstance(fill_color, str) and fill_color not in ('FFFFFFFF', '00000000', 'FFFFFF', '000000'):
                        target.fill = PatternFill(
                            start_color=fill_color, 
                            end_color=fill_color, 
                            fill_type='solid'
                        )
                    else:
                        # No fill (white background)
                        target.fill = PatternFill(fill_type=None)
                else:
                    # No fill (white background)
                    target.fill = PatternFill(fill_type=None)

            self.row_cursor += 1



