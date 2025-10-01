# -*- coding: UTF-8 -*-
"""
WinMerge integration and Excel conversion module
"""
import os
import sys
import subprocess
from pathlib import Path
from typing import Optional, Callable

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .config import config
from .exceptions import WinMergeNotFoundError, ExcelProcessingError, FileProcessingError
from .utils import FileNormalizer, ExcelFormatter, PathManager, clean_output_files
from .diffdetailsheetcreater import DiffDetailSheetCreator
from .common import Timer, logger


def try_close_excel_file(file_path: Path) -> bool:
    """
    Try to close an Excel file if it's open
    
    Returns:
        True if file was closed or not open, False if failed to close
    """
    try:
        import psutil
        
        file_path = str(file_path).lower()
        
        for proc in psutil.process_iter(['pid', 'name', 'open_files']):
            try:
                if proc.info['name'] and 'excel' in proc.info['name'].lower():
                    if proc.info['open_files']:
                        for file in proc.info['open_files']:
                            if file_path in file.path.lower():
                                logger.info(f"Found Excel process (PID: {proc.info['pid']}) with file open")
                                # Don't kill the process, just log it
                                return False
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                continue
        
        return True
        
    except ImportError:
        logger.debug("psutil not available, skipping Excel process check")
        return True
    except Exception as e:
        logger.warning(f"Error checking Excel processes: {e}")
        return True


timer_WMX = Timer("WinMergeXlsx")
class WinMergeXlsx:
    """Main class for WinMerge integration and Excel conversion"""
    
    def __init__(self, base: str, latest: str, output: str = './output.xlsx', 
                 log_callback: Optional[Callable[[str], None]] = None):
        self.base = Path(base).absolute()
        self.latest = Path(latest).absolute()
        self.output = Path(output).absolute()
        self.log_callback = log_callback
        
        self.output_html = self.output.with_suffix('.html')
        self.output_html_files = self.output_html.with_name(self.output_html.stem + '.files')
        
        self.path_manager = PathManager()
        
        # Sheet name to original filename mapping
        self.sheet_name_to_filename = {}
        
        self._validate_inputs()
        self._setup()

    def log(self, message: str) -> None:
        """Log message using callback if available"""
        if self.log_callback:
            self.log_callback(message)
        logger.info(message)

    def generate(self) -> None:
        """Main generation process"""
        timer_WMX.start(memo="generate")
        try:
            self._normalize_files()
            self._generate_html_by_winmerge()
            self._convert_html_to_xlsx()
            DiffDetailSheetCreator(str(self.output), sheet_name_to_filename=self.sheet_name_to_filename).generate()
            self.log("Generation completed successfully")
        except Exception as e:
            logger.error(f"Generation failed: {e}")
            raise
        finally:
            self._cleanup()
            timer_WMX.stop()

    def _validate_inputs(self) -> None:
        """Validate input parameters"""
        if not self.base.exists():
            raise FileProcessingError(f"Base path does not exist: {self.base}")
        if not self.latest.exists():
            raise FileProcessingError(f"Latest path does not exist: {self.latest}")
        
        winmerge_path = Path(config.winmerge.executable_path)
        if not winmerge_path.exists():
            raise WinMergeNotFoundError(f"WinMerge not found at: {winmerge_path}")

    def _setup(self) -> None:
        """Setup application environment"""
        self._clean_output_files()

    def _clean_output_files(self) -> None:
        """Clean existing output files"""
        try:
            clean_output_files(self.output_html, self.output_html_files, self.output)
        except Exception as e:
            # Don't fail if cleanup fails - we'll try to overwrite
            logger.warning(f"File cleanup warning: {e}")
            self.log("Note: Output files may be in use. Will attempt to overwrite.")

    def _generate_html_by_winmerge(self) -> None:
        """Generate HTML report using WinMerge"""
        self.log("Generating HTML report with WinMerge...")
        
        command = config.get_winmerge_command(
            str(self.normalized_base),
            str(self.normalized_latest),
            str(self.output_html)
        )
        
        logger.debug(f"WinMerge command: {' '.join(command)}")
        
        try:
            result = subprocess.run(
                command,
                check=True,
                capture_output=True,
                text=True,
                timeout=300  # 5 minutes timeout
            )
            self.log("WinMerge HTML generation completed")
            if result.stdout:
                logger.debug(f"WinMerge output: {result.stdout}")
        except subprocess.TimeoutExpired:
            raise ExcelProcessingError("WinMerge execution timed out after 5 minutes")
        except subprocess.CalledProcessError as e:
            error_msg = f"WinMerge execution failed: {e}"
            if e.stderr:
                error_msg += f"\nError output: {e.stderr}"
            raise ExcelProcessingError(error_msg)

    def _convert_html_to_xlsx(self) -> None:
        """Convert HTML report to Excel using pure Python (no Excel COM)"""
        timer_WMX.start(memo="convert_html_to_xlsx")
        
        try:
            from src.converters.html_to_excel import HTMLToExcelConverter
            
            self.log("Converting HTML to Excel (no Excel installation required)...")
            
            # Create converter
            converter = HTMLToExcelConverter(log_callback=self.log_callback)
            
            # Convert summary HTML to workbook
            self.wb = converter.convert_summary_html(self.output_html)
            
            # Convert all diff HTML files
            self._convert_diff_html_files(converter)
            
            # Check if file is open in Excel before saving
            if not try_close_excel_file(self.output):
                self.log("Warning: Excel file is currently open. Attempting to save anyway...")
            
            # Save initial version with retry logic
            self._save_workbook_with_retry(self.wb, self.output)
            
            # Now process with openpyxl for formatting
            self._process_with_openpyxl()
            
        except PermissionError as e:
            error_msg = f"Cannot save Excel file. Please close '{self.output.name}' if it's open in Excel."
            logger.error(error_msg, exc_info=True)
            raise ExcelProcessingError(error_msg)
        except Exception as e:
            raise ExcelProcessingError(f"Excel conversion failed: {e}")
        finally:
            timer_WMX.stop()

    def _save_workbook_with_retry(self, wb, output_path: Path, max_retries: int = 5) -> None:
        """
        Save workbook with retry logic for file locks
        Uses temporary file strategy if direct save fails
        
        Args:
            wb: Workbook to save
            output_path: Path to save to
            max_retries: Maximum number of retry attempts
        """
        import time
        import tempfile
        import shutil
        
        # Strategy 1: Direct save with retry
        for attempt in range(max_retries):
            try:
                wb.save(str(output_path))
                logger.info(f"Workbook saved successfully: {output_path}")
                return
            except PermissionError as e:
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 0.5
                    logger.warning(f"File is locked, retrying in {wait_time}s... ({attempt + 1}/{max_retries})")
                    self.log(f"File is locked. Retrying in {wait_time} seconds...")
                    time.sleep(wait_time)
                else:
                    logger.warning(f"Direct save failed after {max_retries} attempts, trying alternative method...")
                    break
            except Exception as e:
                logger.error(f"Unexpected error saving workbook: {e}", exc_info=True)
                raise
        
        # Strategy 2: Save to temp file, then attempt to replace
        try:
            self.log("Attempting alternative save method (temporary file)...")
            
            # Create temp file in same directory for atomic rename
            temp_fd, temp_path = tempfile.mkstemp(
                suffix='.xlsx',
                prefix='~temp_',
                dir=output_path.parent
            )
            os.close(temp_fd)  # Close file descriptor
            temp_path = Path(temp_path)
            
            # Save to temp file
            wb.save(str(temp_path))
            logger.info(f"Saved to temporary file: {temp_path}")
            
            # Try to replace original file
            for attempt in range(3):
                try:
                    # On Windows, need to delete first if exists
                    if output_path.exists():
                        output_path.unlink()
                    
                    # Rename temp to target
                    shutil.move(str(temp_path), str(output_path))
                    logger.info(f"Successfully replaced file: {output_path}")
                    self.log(f"Excel file saved successfully: {output_path}")
                    return
                    
                except PermissionError:
                    if attempt < 2:
                        logger.warning(f"Replace attempt {attempt + 1} failed, retrying...")
                        time.sleep(1.0)
                    else:
                        # Save with timestamp instead
                        raise
            
        except Exception as e:
            # Strategy 3: Save with timestamp suffix
            logger.warning(f"Alternative save failed: {e}")
            self._save_with_timestamp(wb, output_path)
    
    def _save_with_timestamp(self, wb, output_path: Path) -> None:
        """Save workbook with timestamp suffix as last resort"""
        from datetime import datetime
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = output_path.stem + f"_{timestamp}" + output_path.suffix
        new_path = output_path.parent / new_name
        
        try:
            wb.save(str(new_path))
            logger.info(f"Saved to alternative file: {new_path}")
            self.log(f"Original file is locked. Saved as: {new_name}")
            self.log(f"Please close Excel and rename the file manually if needed.")
            
            # Update output path for subsequent operations
            self.output = new_path
            
        except Exception as e:
            logger.error(f"All save strategies failed: {e}", exc_info=True)
            raise ExcelProcessingError(
                f"Cannot save Excel file. All attempts failed.\n"
                f"Please close '{output_path.name}' in Excel and try again."
            )

    def _extract_filename_from_stem(self, stem: str) -> str:
        """
        Extract actual filename from HTML file stem.
        
        HTML files from WinMerge use pattern: folder1_folder2_..._filename
        We need to identify where the path ends and filename begins.
        
        Strategy:
        1. Look for file extension (contains '.')
        2. Scan backwards from extension to find filename pattern
        3. Consider common folder names (modules, ctrl, tool, etc.) as path components
        
        Examples:
            "modcommon_modules_SPM_SEQ_spm_tbl_rinse.t" -> "spm_tbl_rinse.t"
            "sc2stb_ctrl_v4_LINK_CTRLA_SP3_2.l" -> "CTRLA_SP3_2.l"
            "sc2stb_modules_SPM_tool_gui_Gui_Res_spm_en-US.rc" -> "spm_en-US.rc"
            "sc2stb_etc_cset_file.SP3" -> "cset_file.SP3"
        """
        parts = stem.split('_')
        
        # Common folder names that are NOT part of filenames
        folder_keywords = {
            'modules', 'ctrl', 'tool', 'gui', 'etc', 'src', 'include',
            'lib', 'bin', 'obj', 'SEQ', 'u', 'v4', 'LINK', 'Res',
            'modcommon', 'sc2stb', 'Gui'
        }
        
        # Find the extension position (last part with '.')
        ext_index = -1
        for i in range(len(parts) - 1, -1, -1):
            if '.' in parts[i]:
                ext_index = i
                break
        
        if ext_index == -1:
            # No extension found, return last part
            return parts[-1]
        
        # Scan backwards from extension to find start of filename
        # A filename typically starts after a known folder keyword
        filename_start = 0
        
        for i in range(ext_index, -1, -1):
            part = parts[i]
            
            # Check if this part is a folder keyword
            if part in folder_keywords:
                filename_start = i + 1
                break
            
            # If we reach the beginning, include from start
            if i == 0:
                filename_start = 0
        
        # Special case: if filename_start is too close to extension, 
        # take at least 2 parts before extension
        if ext_index - filename_start < 1:
            filename_start = max(0, ext_index - 1)
        
        # Build filename from identified parts
        filename_parts = parts[filename_start:ext_index + 1]
        filename = '_'.join(filename_parts)
        
        return filename

    def _convert_diff_html_files(self, converter) -> None:
        """Convert all diff HTML files to Excel sheets"""
        html_files = sorted(self.output_html_files.glob('**/*.html'))
        self.log(f"Processing {len(html_files)} diff HTML files...")
        
        # Track sheet names to handle duplicates
        used_sheet_names = set()
        
        for count, html_file in enumerate(html_files, start=1):
            # Generate sheet name from file
            original_name = html_file.stem
            
            # Extract filename from path-like stem more reliably
            # HTML file names follow pattern: folder1_folder2_..._filename.ext
            # We need to find where the path ends and filename begins
            filename = self._extract_filename_from_stem(original_name)
            
            logger.info(f"[DEBUG] HTML file: {html_file.name}")
            logger.info(f"[DEBUG] HTML stem (original): '{original_name}'")
            logger.info(f"[DEBUG] Extracted filename: '{filename}'")
            
            # Use filename as sheet name (more readable than full path)
            sheet_name = filename
            
            # Handle Excel's 31 character limit for sheet names
            if len(sheet_name) > 31:
                # Truncate and add counter
                sheet_name = sheet_name[:28] + f"_{count}"
                logger.info(f"[DEBUG] Sheet name too long, truncated: '{filename}' -> '{sheet_name}'")
            
            # Handle duplicate sheet names
            original_sheet_name = sheet_name
            suffix = 1
            while sheet_name in used_sheet_names:
                # If duplicate, add/increment suffix
                if len(original_sheet_name) > 28:
                    sheet_name = original_sheet_name[:27] + f"_{suffix}"
                else:
                    sheet_name = f"{original_sheet_name}_{suffix}"
                suffix += 1
                logger.info(f"[DEBUG] Duplicate sheet name, adjusted: '{original_sheet_name}' -> '{sheet_name}'")
            
            used_sheet_names.add(sheet_name)
            
            # Store mapping from sheet name to actual filename
            # (in this case, they should be the same or very similar)
            self.sheet_name_to_filename[sheet_name] = filename
            logger.info(f"[DEBUG] Final sheet name: '{sheet_name}'")
            
            converter.convert_html_file(html_file, self.wb, sheet_name)
            
            if count % 10 == 0:
                self.log(f"Processed {count}/{len(html_files)} files...")

    def _process_with_openpyxl(self) -> None:
        """Process Excel file using openpyxl for final formatting"""
        self.log("Applying final formatting with openpyxl...")
        
        # Reload the workbook to ensure proper processing
        self.wb = load_workbook(filename=self.output)
        self.summary_ws = self.wb.worksheets[config.excel.summary_ws_num - 1]
        
        self._format_summary_sheet()
        self._format_diff_sheets()
        self._save_workbook()

    def _format_summary_sheet(self) -> None:
        """Format summary sheet"""
        logger.info("Formatting summary sheet")
        
        for row in self.summary_ws.iter_rows(min_row=config.excel.summary_start_row):
            name_cell = row[config.excel.summary_name_col_index]
            if not name_cell.value:
                break
                
            name_cell.hyperlink = f"{name_cell.value}!{config.excel.home_position}"
            folder_cell = row[config.excel.summary_folder_col_index]
            
            if folder_cell.value:
                self._rename_html_files(name_cell.value, folder_cell.value)

    def _rename_html_files(self, name: str, folder: str) -> None:
        """Rename HTML files for proper linking"""
        sheet_name = folder.replace('\\', '_') + '_' + name
        src = self.output_html_files / f"{sheet_name}.html"
        dst = self.output_html_files / f"{name}.html"
        
        if src.exists():
            os.rename(src, dst)

    def _format_diff_sheets(self) -> None:
        """Format diff sheets"""
        logger.info("Formatting diff sheets")
        
        for ws in self.wb.worksheets[config.excel.diff_start_row - 1:]:
            self._remove_hyperlinks_from_line_numbers(ws)
            ExcelFormatter.set_worksheet_format(ws)

    def _remove_hyperlinks_from_line_numbers(self, ws) -> None:
        """Remove hyperlinks from line number columns"""
        for fmt in config.diff_formats['no']:
            col_letter = fmt['col']
            
            for row in range(config.excel.diff_start_row, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.hyperlink = None
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                cell.font = Font(size=12)

    def _save_workbook(self) -> None:
        """Save the workbook with retry logic"""
        self._save_workbook_with_retry(self.wb, self.output)
        self.log(f"Excel file saved: {self.output}")

    def _normalize_files(self) -> None:
        """Normalize versioned files and copy to temporary directories"""
        logger.info("Normalizing files")
        
        self.temp_base = self.path_manager.create_temp_dir()
        self.temp_latest = self.path_manager.create_temp_dir()

        FileNormalizer.copy_and_normalize(
            self.base, self.temp_base, 
            lambda msg: self.log(f"Base: {msg}")
        )
        FileNormalizer.copy_and_normalize(
            self.latest, self.temp_latest,
            lambda msg: self.log(f"Latest: {msg}")
        )

        self.normalized_base = self.temp_base
        self.normalized_latest = self.temp_latest

    def _cleanup(self) -> None:
        """Clean up resources"""
        self.path_manager.cleanup()
        logger.info("Cleanup completed")
