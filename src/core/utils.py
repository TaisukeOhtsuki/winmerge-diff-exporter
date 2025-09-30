# -*- coding: UTF-8 -*-
"""
Utility functions for file operations and formatting
"""
import os
import shutil
import tempfile
from pathlib import Path
from typing import List, Optional, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import column_index_from_string

from .config import config
from .exceptions import FileProcessingError
from .common import logger


class FileNormalizer:
    """Handles file normalization operations"""
    
    @staticmethod
    def normalize_filename(filename: str) -> str:
        """
        Normalize file name by removing version suffixes
        
        Example: io.h.202334 -> io.h
        
        Args:
            filename: Original filename
            
        Returns:
            Normalized filename
        """
        logger.debug(f"Normalizing filename: {filename}")
        parts = filename.split('.')
        # Remove trailing numeric extension if present (e.g., version numbers)
        return '.'.join(parts[:-1]) if len(parts) >= 3 and parts[-1].isdigit() else filename
    
    @staticmethod
    def copy_and_normalize(src: Path, dest: Path, progress_callback: Optional[Callable] = None) -> None:
        """Copy and normalize files from source to destination"""
        logger.info(f"Copying and normalizing: {src} -> {dest}")
        
        if not src.exists():
            raise FileProcessingError(f"Source path does not exist: {src}")
        
        # Determine files to process
        if src.is_file():
            files = [src]
            src_base = src.parent
        elif src.is_dir():
            files = [f for f in src.rglob('*') if f.is_file()]
            src_base = src
        else:
            raise FileProcessingError(f"Source path is neither file nor directory: {src}")
        
        logger.info(f"Found {len(files)} files to process")
        
        def copy_file(file: Path) -> None:
            try:
                norm_name = FileNormalizer.normalize_filename(file.name)
                
                if dest.is_file() or (len(files) == 1 and not dest.exists()):
                    target_path = dest if dest.suffix else dest / norm_name
                else:
                    rel_path = file.relative_to(src_base).parent / norm_name
                    target_path = dest / rel_path
                
                target_path.parent.mkdir(parents=True, exist_ok=True)
                logger.info(f"Copying: {file} -> {target_path}")
                shutil.copy2(file, target_path)
                
                if progress_callback:
                    progress_callback(f"Copied: {file.name}")
                    
            except Exception as e:
                logger.error(f"Failed to copy file {file}: {e}")
                raise FileProcessingError(f"Failed to copy file {file}: {e}")
        
        with ThreadPoolExecutor() as executor:
            executor.map(copy_file, files)


class ExcelFormatter:
    """Handles Excel worksheet formatting"""
    
    @staticmethod
    def set_worksheet_format(ws, end_row: Optional[int] = None) -> None:
        """Apply formatting to worksheet"""
        if end_row is None:
            end_row = ws.max_row
        
        for format_type, formats in config.diff_formats.items():
            for fmt in formats:
                col_letter = fmt['col']
                
                if 'width' in fmt:
                    ws.column_dimensions[col_letter].width = fmt['width']
                
                if 'font' in fmt:
                    for cell in ws[col_letter]:
                        cell.font = Font(name=fmt['font'])
                
                if 'comment' in fmt:
                    ExcelFormatter._set_extra_table(ws, fmt, end_row)
                
                if 'header' in fmt:
                    ExcelFormatter._set_header(ws, fmt)
        
        # Apply borders to all cells in the worksheet for better visibility
        ExcelFormatter._apply_borders_to_sheet(ws, end_row)
    
    @staticmethod
    def _set_header(ws, fmt: dict) -> None:
        """Set header for column"""
        cell = ws[fmt['col'] + '1']
        cell.value = fmt['header']
        cell.font = Font(bold=True)
    
    @staticmethod
    def _set_extra_table(ws, fmt: dict, end_row: Optional[int] = None) -> None:
        """Set extra table formatting"""
        if end_row is None:
            end_row = ws.max_row
        
        # Set header
        header_cell = ws[fmt['col'] + '1']
        header_cell.value = fmt['comment']
        header_cell.alignment = Alignment(vertical='center', horizontal='center')
        header_cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        
        # Set borders (vertical only for cleaner look)
        vertical_border = Border(
            left=Side(style='thin', color='C0C0C0'),
            right=Side(style='thin', color='C0C0C0')
            # No top/bottom borders
        )
        
        for i in range(1, end_row + 1):
            cell = ws[fmt['col'] + str(i)]
            cell.border = vertical_border
        
        # Fill empty cells
        for i in range(config.excel.diff_start_row, end_row + 1):
            code_col = config.diff_formats['code'][0]['col']
            code_cell = ws[code_col + str(i)]
            if code_cell.fill.start_color.rgb in ('FFFFFFFF', '00000000'):
                target_cell = ws[fmt['col'] + str(i)]
                target_cell.value = '-'
                target_cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    @staticmethod
    def _apply_borders_to_sheet(ws, end_row: int) -> None:
        """Apply borders to all cells in the sheet for better visibility"""
        # Create border style with only vertical lines (no horizontal lines for cleaner look)
        vertical_only_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0')
            # No top/bottom borders for cleaner horizontal view
        )
        
        # Get all columns that have data
        max_col = ws.max_column
        
        # Apply borders to all cells in the data range
        for row in range(1, end_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = vertical_only_border


class PathManager:
    """Manages temporary directories and file paths"""
    
    def __init__(self):
        self.temp_dirs: List[Path] = []
    
    def create_temp_dir(self) -> Path:
        """Create a temporary directory"""
        temp_dir = Path(tempfile.mkdtemp())
        self.temp_dirs.append(temp_dir)
        return temp_dir
    
    def cleanup(self) -> None:
        """Clean up all temporary directories"""
        for temp_dir in self.temp_dirs:
            if temp_dir.exists():
                try:
                    shutil.rmtree(temp_dir)
                    logger.info(f"Cleaned up temporary directory: {temp_dir}")
                except Exception as e:
                    logger.warning(f"Failed to clean up {temp_dir}: {e}")
        self.temp_dirs.clear()
    
    def __del__(self):
        self.cleanup()


def clean_output_files(*paths: Path) -> None:
    """Clean up output files if they exist"""
    import time
    
    for path in paths:
        try:
            if path.exists():
                if path.is_dir():
                    shutil.rmtree(path)
                else:
                    # Try multiple times with delay for locked files
                    max_retries = 3
                    for attempt in range(max_retries):
                        try:
                            path.unlink()
                            logger.info(f"Cleaned up: {path}")
                            break
                        except PermissionError:
                            if attempt < max_retries - 1:
                                logger.warning(f"File locked, retrying... ({attempt + 1}/{max_retries})")
                                time.sleep(0.5)
                            else:
                                logger.warning(f"Could not delete {path} (file may be open in Excel). Will overwrite instead.")
        except PermissionError:
            logger.warning(f'Permission denied for: {path}. File may be open in Excel. Will attempt to overwrite.')
        except Exception as e:
            logger.warning(f"Failed to clean up {path}: {e}")
