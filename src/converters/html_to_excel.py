# -*- coding: utf-8 -*-
"""
HTML to Excel converter without COM dependencies
"""
from pathlib import Path
from typing import Optional, Callable
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.core.common import logger


class HTMLToExcelConverter:
    """Convert HTML tables to Excel without using Excel COM"""
    
    def __init__(self, log_callback: Optional[Callable[[str], None]] = None):
        self.log_callback = log_callback
    
    def log(self, message: str) -> None:
        """Log message"""
        if self.log_callback:
            self.log_callback(message)
        logger.info(message)
    
    def convert_html_file(self, html_path: Path, wb: Workbook, sheet_name: str) -> None:
        """
        Convert single HTML file to Excel worksheet
        
        Args:
            html_path: Path to HTML file
            wb: Workbook to add sheet to
            sheet_name: Name for the new sheet
        """
        try:
            with open(html_path, 'r', encoding='utf-8', errors='ignore') as f:
                html_content = f.read()
            
            soup = BeautifulSoup(html_content, 'html.parser')
            tables = soup.find_all('table')
            
            if not tables:
                logger.warning(f"No tables found in {html_path}")
                return
            
            # Create new worksheet
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit: 31 chars
            
            # Process the main table (usually the first one)
            table = tables[0]
            self._convert_table_to_sheet(table, ws)
            
            self.log(f"Converted: {sheet_name}")
            
        except Exception as e:
            logger.error(f"Failed to convert {html_path}: {e}", exc_info=True)
    
    def _convert_table_to_sheet(self, table, ws) -> None:
        """Convert HTML table to Excel worksheet"""
        row_idx = 1
        
        # Process table rows
        for tr in table.find_all('tr'):
            col_idx = 1
            
            # Process cells (th or td)
            for cell in tr.find_all(['th', 'td']):
                # Get cell text
                cell_text = cell.get_text(strip=True)
                
                # Remove "." from line number columns (columns 1 and 2)
                if col_idx in [1, 2] and cell_text == '.':
                    cell_text = ''
                
                # Write to Excel
                excel_cell = ws.cell(row=row_idx, column=col_idx, value=cell_text)
                
                # Apply basic styling based on HTML attributes
                self._apply_cell_styling(excel_cell, cell)
                
                # Handle colspan and rowspan
                colspan = int(cell.get('colspan', 1))
                rowspan = int(cell.get('rowspan', 1))
                
                if colspan > 1 or rowspan > 1:
                    ws.merge_cells(
                        start_row=row_idx,
                        start_column=col_idx,
                        end_row=row_idx + rowspan - 1,
                        end_column=col_idx + colspan - 1
                    )
                
                col_idx += colspan
            
            row_idx += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _apply_cell_styling(self, excel_cell, html_cell) -> None:
        """Apply styling from HTML cell to Excel cell"""
        # Check for background color
        style = html_cell.get('style', '')
        bgcolor = html_cell.get('bgcolor', '')
        
        # Parse background color
        if 'background-color' in style or bgcolor:
            color = self._parse_color(style, bgcolor)
            if color:
                excel_cell.fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type='solid'
                )
        
        # Check if it's a header cell
        if html_cell.name == 'th':
            excel_cell.font = Font(bold=True, size=11)
            excel_cell.alignment = Alignment(horizontal='center', vertical='center')
            if not excel_cell.fill.start_color.rgb:
                excel_cell.fill = PatternFill(
                    start_color='DDDDDD',
                    end_color='DDDDDD',
                    fill_type='solid'
                )
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        excel_cell.border = thin_border
    
    def _parse_color(self, style: str, bgcolor: str) -> Optional[str]:
        """Parse color from HTML style or bgcolor attribute"""
        # Try bgcolor attribute first
        if bgcolor:
            return self._normalize_color(bgcolor)
        
        # Parse from style attribute
        if 'background-color' in style:
            # Extract color value
            parts = style.split('background-color')
            if len(parts) > 1:
                color_part = parts[1].split(':')[1].split(';')[0].strip()
                return self._normalize_color(color_part)
        
        return None
    
    def _normalize_color(self, color: str) -> Optional[str]:
        """Normalize HTML color to Excel RGB format"""
        color = color.strip().upper()
        
        # Remove # if present
        if color.startswith('#'):
            color = color[1:]
        
        # Convert 3-char hex to 6-char
        if len(color) == 3:
            color = ''.join([c*2 for c in color])
        
        # Add FF prefix for full opacity if not present
        if len(color) == 6:
            color = 'FF' + color
        
        return color if len(color) == 8 else None
    
    def _auto_adjust_columns(self, ws, max_width: int = 100) -> None:
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def convert_summary_html(self, html_path: Path) -> Workbook:
        """
        Convert summary HTML file to a new Workbook
        
        Args:
            html_path: Path to summary HTML file
            
        Returns:
            New Workbook with summary sheet
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        try:
            with open(html_path, 'r', encoding='utf-8', errors='ignore') as f:
                html_content = f.read()
            
            soup = BeautifulSoup(html_content, 'html.parser')
            tables = soup.find_all('table')
            
            if tables:
                self._convert_table_to_sheet(tables[0], ws)
                self.log("Converted summary HTML to Excel")
            else:
                logger.warning(f"No tables found in summary HTML: {html_path}")
        
        except Exception as e:
            logger.error(f"Failed to convert summary HTML: {e}", exc_info=True)
        
        return wb
