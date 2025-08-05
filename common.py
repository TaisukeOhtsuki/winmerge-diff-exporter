# -*- coding: UTF-8 -*-
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
xlUp = -4162
xlOpenXMLWorkbook = 51
xlCenter = -4108
xlContinuous = 1

SUMMARY_WS_NUM = 1        # 一覧シートのワークシート番号
SUMMARY_START_ROW = 6     # 一覧シートの表の開始行
SUMMARY_NAME_COL_INDEX = 1
SUMMARY_FOLDER_COL_INDEX = 2

HOME_POSITION = 'A1'  # ホームポジション

DIFF_START_ROW = 2                                            # 差分シートの開始行
DIFF_ZOOM_RATIO = 85                                          # 差分シートのズームの倍率
DIFF_FORMATS = {                                              # 差分シートの書式設定
    'no': [                                                   # 行番号列
        {'col': 'A', 'width': 5},                             # 左側
        {'col': 'C', 'width': 5},                             # 右側
    ],
    'code': [                                                 # ソースコード列
        {'col': 'B', 'width': 100, 'font': 'ＭＳ ゴシック','header':'更新前'},  # 左側
        {'col': 'D', 'width': 100, 'font': 'ＭＳ ゴシック','header':'更新後'},  # 右側
    ],
    'extra': [                                                # 追加列
        {'col': 'E', 'width': 60, 'comment': 'コメント'},
    ],
}


def common_set_format(ws, row=None):
    for key in DIFF_FORMATS.keys():
        for f in DIFF_FORMATS[key]:
            col_letter = f['col']
            col_index = column_index_from_string(col_letter)
            if 'width' in f:
                ws.column_dimensions[col_letter].width = f['width']
            if 'font' in f:
                for cell in ws[col_letter]:
                    cell.font = Font(name=f['font'])
            if 'comment' in f:
                common_set_extra_table(ws, f, end_row=row)
            if 'header' in f:
                common_set_header(ws, f)

def common_set_header(ws, f):
    cell = ws[f['col'] + '1']
    cell.value = f['header']
    cell.font = Font(bold=True)

def common_set_extra_table(ws, f, end_row=None):
    if end_row is None:
        end_row = ws.max_row

    header_cell = ws[f['col'] + '1']
    header_cell.value = f['comment']
    header_cell.alignment = Alignment(vertical='center', horizontal='center')
    header_cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for i in range(1, end_row + 1):
        cell = ws[f['col'] + str(i)]
        cell.border = thin_border

    for i in range(DIFF_START_ROW, end_row + 1):
        code_col = DIFF_FORMATS['code'][0]['col']
        code_cell = ws[code_col + str(i)]
        if code_cell.fill.start_color.rgb in ('FFFFFFFF', '00000000'):
            target_cell = ws[f['col'] + str(i)]
            target_cell.value = '-'
            target_cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

import time

class Timer:
    def __init__(self, label: str = "処理"):
        self.label = label
        self._sessions = []
        self._memos = []
    def start(self, memo:str = None) -> None:
        self._sessions.append({'start': time.time(), 'end': None})
        indent = '_' * (len(self._sessions) * 2)
        self._memos.append(memo)
        print(f"{indent}{self.label} {memo}を開始します...")

    def stop(self) -> None:
        if not self._sessions or self._sessions[-1]['end'] is not None:
            print(f"{self.label}はまだ開始されていません。")
            return

        self._sessions[-1]['end'] = time.time()
        elapsed = self._sessions[-1]['end'] - self._sessions[-1]['start']
        indent = '_' * (len(self._sessions) * 2)
        memo = self._memos[-1]
        print(f"{indent}{self.label} {memo}が完了しました。経過時間: {elapsed:.2f} 秒")

        # セッションを削除（完了したもの）
        self._sessions.pop()
        self._memos.pop()

    def elapsed_all(self) -> float:
        return sum(
            (s['end'] or time.time()) - s['start']
            for s in self._sessions
        )



import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Optional

class Logger:
    def __init__(
        self,
        name: str = "logger",
        level: int = logging.INFO,
        log_file: Optional[Path] = None,
        max_bytes: int = 1_000_000,  # 1MB
        backup_count: int = 3
    ):
        self.logger = logging.getLogger(name)
        self.logger.setLevel(level)
        self.logger.propagate = False

        if not self.logger.handlers:
            formatter = logging.Formatter(
                "%(asctime)s | %(levelname)s | %(name)s | %(message)s",
                datefmt="%Y-%m-%d %H:%M:%S"
            )

            # Console handler
            console_handler = logging.StreamHandler()
            console_handler.setFormatter(formatter)
            self.logger.addHandler(console_handler)

            # File handler with rotation
            if log_file:
                log_file = Path(log_file)
                log_file.parent.mkdir(parents=True, exist_ok=True)
                file_handler = RotatingFileHandler(
                    log_file, maxBytes=max_bytes, backupCount=backup_count, encoding="utf-8"
                )
                file_handler.setFormatter(formatter)
                self.logger.addHandler(file_handler)

    def debug(self, message: str) -> None:
        self.logger.debug(message)

    def info(self, message: str) -> None:
        self.logger.info(message)

    def warning(self, message: str) -> None:
        self.logger.warning(message)

    def error(self, message: str) -> None:
        self.logger.error(message)

    def critical(self, message: str) -> None:
        self.logger.critical(message)

logger = Logger()