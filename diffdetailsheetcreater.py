# -*- coding: UTF-8 -*-
from openpyxl.styles import Font,Color, PatternFill,Font, PatternFill
from openpyxl import load_workbook
from common import *

timer_DDSC = Timer("DiffDetailSheetCreator")

class DiffDetailSheetCreator:
    def __init__(self, file_path: str, start_index: int = 2, context_lines: int = 4):
        self.file_path = file_path
        self.wb = load_workbook(self.file_path)
        self.start_index = start_index
        self.context_lines = context_lines
        self.detail_ws = self.wb.create_sheet(index=0, title='compare')
        
        self.row_cursor = 2

    def generate(self) -> None:
        timer_DDSC.start(memo="generate")
        for ws in self.wb.worksheets[self.start_index:]:
            file_name = self._extract_filename(ws.title)
            self._write_filename_label(file_name)
            self._process_sheet(ws)
        common_set_format(self.detail_ws)
        self.wb.save(self.file_path)
        timer_DDSC.stop()

    def _extract_filename(self, sheet_name: str) -> str:
        return sheet_name.replace('_', '\\')

    def _write_filename_label(self, file_name: str) -> None:
        cell = self.detail_ws.cell(row=self.row_cursor, column=1)
        cell.value = file_name
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')
        self.row_cursor += 1

    def _process_sheet(self, ws) -> None:
        max_row = self._get_max_colored_row(ws)
        diff_rows = self._detect_diff_rows(ws, max_row)
        blocks = self._merge_diff_blocks(diff_rows)

        for block_start, block_end in blocks:
            self._copy_block(ws, block_start, block_end)
            self.row_cursor += 2  # Add spacing between blocks

        self.row_cursor += 2  # Add spacing after processing each sheet

    def _get_max_colored_row(self, ws) -> int:
        for row in reversed(range(1, ws.max_row + 1)):
            cell = ws.cell(row=row, column=1)
            if cell.fill.start_color.rgb not in ('FFFFFFFF', '00000000'):
                return row
        print("No colored rows found; using max_row")
        return ws.max_row

    def _detect_diff_rows(self, ws, max_row: int) -> list[int]:
        timer_DDSC.start(memo="detect_diff_rows")
        diff_rows = set()
        yellow = 'FFC0C0C0'  # RGB for yellow in openpyxl
        for row in range(DIFF_START_ROW, max_row + 1):
            for col_letter in [item['col'] for item in DIFF_FORMATS['code']]:
                col = column_index_from_string(col_letter)
                cell = ws.cell(row=row, column=col)
                if cell.fill.start_color.rgb == yellow:
                    diff_rows.add(row)
                    break
        timer_DDSC.stop()
        return sorted(diff_rows)

    def _merge_diff_blocks(self, diff_rows: list[int]) -> list[tuple[int, int]]:
        timer_DDSC.start(memo="merge_diff_blocks")
        if not diff_rows:
            return []

        ranges = [(max(row - self.context_lines, DIFF_START_ROW), row + self.context_lines) for row in diff_rows]
        merged = []
        current_start, current_end = ranges[0]

        for start, end in ranges[1:]:
            if start <= current_end + 1:
                current_end = max(current_end, end)
            else:
                merged.append((current_start, current_end))
                current_start, current_end = start, end

        merged.append((current_start, current_end))
        timer_DDSC.stop()
        return merged

    def _copy_block(self, ws, start_row: int, end_row: int) -> None:
        for row in range(start_row, end_row + 1):
            for src_col, dst_col in zip([1, 2, 3, 4], range(1, 7)):
                source = ws.cell(row=row, column=src_col)
                target = self.detail_ws.cell(row=self.row_cursor, column=dst_col)
                target.value = source.value

                # Copy font color
                font_color = getattr(source.font.color, 'rgb', None)
                target.font = Font(color=Color(rgb=font_color)) if isinstance(font_color, str) else Font()

                # Copy fill color
                fill_color = getattr(source.fill.start_color, 'rgb', 'FFFFFFFF')
                target.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            self.row_cursor += 1



