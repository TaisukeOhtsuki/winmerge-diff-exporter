# -*- coding: UTF-8 -*-
import os
import sys
import shutil
import tempfile
import subprocess
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from diffdetailsheetcreater import DiffDetailSheetCreator

from common import *

timer_WMX = Timer("main")

WINMERGE_EXE = r'C:\Program Files\WinMerge\WinMergeU.exe'  # WinMergeへのパス
WINMERGE_OPTIONS = [
    '/minimize',                           # ウィンドウ最小化で起動
    '/noninteractive',                     # レポート出力後に終了
    '/cfg',
    'Settings/DirViewExpandSubdirs=1',     # 自動的にサブフォルダーを展開する
    '/cfg',
    'ReportFiles/ReportType=2',            # シンプルなHTML形式
    '/cfg',
    'ReportFiles/IncludeFileCmpReport=1',  # ファイル比較レポートを含める
    '/cfg',
    'Settings/ViewLineNumbers=1',            # 行番号を表示する
    '/cfg',
    'Settings/IgnoreEol=1',                #改行文字違いを無視
    '/r',                                  # すべてのサブフォルダ内のすべてのファイルを比較
    '/u',                                  # 最近使用した項目リストに追加しない
    '/or',                                 # レポートを出力
]
class WinMergeXlsx:
    def __init__(self, base: str, latest: str, output: str = './output.xlsx', log_callback=None):
        self.base = Path(base).absolute()
        self.latest = Path(latest).absolute()
        self.output = Path(output).absolute()
        self.log_callback = log_callback

        self.output_html = self.output.with_suffix('.html')
        self.output_html_files = self.output_html.with_name(self.output_html.stem + '.files')

        self._setup()

    def log(self, message: str) -> None:
        if self.log_callback:
            self.log_callback(message)

    def generate(self) -> None:
        timer_WMX.start(memo="generate")
        self._normalize_files()
        self._generate_html_by_winmerge()
        self._convert_html_to_xlsx()
        DiffDetailSheetCreator(str(self.output)).generate()
        timer_WMX.stop()
    def _setup(self) -> None:
        self._setup_excel_application()
        self._setup_output_files()

    def _setup_excel_application(self) -> None:
        try:
            import win32com.client
            if win32com.client.GetObject(Class='Excel.Application'):
                self.__message_and_exit('Please close Excel before running this process.')
        except Exception:
            logger.info('Unable to conduct setup_excel_application.')
            pass

    def _setup_output_files(self) -> None:
        for path in [self.output_html, self.output_html_files, self.output]:
            try:
                if path.exists():
                    if path.is_dir():
                        shutil.rmtree(path)
                    else:
                        path.unlink()
            except PermissionError:
                self.__message_and_exit(f'Permission denied for: {path}')

    def __message_and_exit(self, message: str) -> None:
        self.log(f'\nError : {message}')
        logger.error(f'\nError : {message}')
        sys.exit(-1)

    def _generate_html_by_winmerge(self) -> None:
        command = [
            WINMERGE_EXE,
            str(self.normalized_base),
            str(self.normalized_latest),
            *WINMERGE_OPTIONS,
            str(self.output_html),
        ]
        subprocess.run(command)

    def _convert_html_to_xlsx(self) -> None:
        timer_WMX.start(memo="convert_html_to_xlsx")
        try:
            import win32com.client
            self.excel = win32com.client.Dispatch('Excel.Application')
            self.wb_com = self.excel.Workbooks.Open(str(self.output_html))
            self.summary_ws_com = self.wb_com.Worksheets(SUMMARY_WS_NUM)

            self._copy_html_files_with_com()
            self.wb_com.SaveAs(str(self.output), FileFormat=xlOpenXMLWorkbook)
        finally:
            self.excel.Quit()

        self._load_workbook_with_openpyxl()
        self._format_summary_sheet_with_openpyxl()
        self._format_diff_sheets_with_openpyxl()
        self._save_book_with_openpyxl()

        timer_WMX.stop()

    def _copy_html_files_with_com(self) -> None:
        timer_WMX.start("copy_html_files_with_com")
        for count, html in enumerate(self.output_html_files.glob('**/*.html'), start=1):
            diff_wb = self.excel.Workbooks.Open(str(html))
            diff_ws = diff_wb.Worksheets(1)
            diff_ws.Copy(Before=None, After=self.wb_com.Worksheets(count))
        timer_WMX.stop()
    def _load_workbook_with_openpyxl(self) -> None:
        logger.info("load_workbook_with_openpyxl")
        self.wb = load_workbook(filename=self.output)
        self.summary_ws = self.wb.worksheets[SUMMARY_WS_NUM - 1]

    def _format_summary_sheet_with_openpyxl(self) -> None:
        logger.info("format_summary_sheet_with_openpyxl")
        for row in self.summary_ws.iter_rows(min_row=SUMMARY_START_ROW):
            name_cell = row[SUMMARY_NAME_COL_INDEX]
            if not name_cell.value:
                break
            name_cell.hyperlink = f"{name_cell.value}!{HOME_POSITION}"
            folder_cell = row[SUMMARY_FOLDER_COL_INDEX]
            if folder_cell.value:
                self._rename_html_files(name_cell.value, folder_cell.value)

    def _rename_html_files(self, name: str, folder: str) -> None:
        sheet_name = folder.replace('\\', '_') + '_' + name
        src = self.output_html_files / f"{sheet_name}.html"
        dst = self.output_html_files / f"{name}.html"
        os.rename(src, dst)

    def _format_diff_sheets_with_openpyxl(self) -> None:
        logger.info("format_diff_sheets_with_openpyxl")
        for i, ws in enumerate(self.wb.worksheets[DIFF_START_ROW - 1:], start=DIFF_START_ROW):
            self._remove_hyperlink_from_no(ws)
            common_set_format(ws)

    def _remove_hyperlink_from_no(self, ws) -> None:
        logger.info("remove_hyperlink_from_no")
        for f in DIFF_FORMATS['no']:
            col_letter = f['col']
            for row in range(DIFF_START_ROW, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.hyperlink = None
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                cell.font = Font(size=12)

    def _save_book_with_openpyxl(self) -> None:
        self.wb.save(str(self.output))

    def _normalize_files(self) -> None:
        logger.info("normalize_files")
        """Normalize versioned files and copy them to temporary directories."""
        self.temp_base = Path(tempfile.mkdtemp())
        self.temp_latest = Path(tempfile.mkdtemp())

        self._copy_and_normalize(self.base, self.temp_base)
        self._copy_and_normalize(self.latest, self.temp_latest)

        self.normalized_base = self.temp_base
        self.normalized_latest = self.temp_latest

    def _copy_and_normalize(self, src: Path, dest: Path) -> None:
        logger.info("copy_and_normalize")

        # file or dir 
        if src.is_file():
            files = [src]
            src_base = src.parent
        elif src.is_dir():
            files = [f for f in src.rglob('*') if f.is_file()]
            src_base = src
        else:
            raise ValueError(f"src path is neither file nor directory: {src}")

        logger.info(f"files: {files}")

        def copy_file(file: Path) -> None:
            norm_name = self._normalize_filename(file.name)

            if dest.is_file() or (len(files) == 1 and not dest.exists()):
                # a file → file copy
                target_path = dest if dest.suffix else dest / norm_name
            else:
                # copy to folder
                rel_path = file.relative_to(src_base).parent / norm_name
                target_path = dest / rel_path

            target_path.parent.mkdir(parents=True, exist_ok=True)
            logger.info(f"copying file {file} -> {target_path}")
            shutil.copy2(file, target_path)

        with ThreadPoolExecutor() as executor:
            executor.map(copy_file, files)

    def _normalize_filename(self, filename: str) -> str:
        """Normalize file name. Example: io.h.202334 -> io.h"""
        logger.info(f"normalizing {filename}")
        parts = filename.split('.')
        return '.'.join(parts[:-1]) if len(parts) >= 3 and parts[-1].isdigit() else filename
